"""Парсинг и reconcile списков пассажиров (ТЗ §11)."""
import csv
import io
import re
from datetime import date, datetime

DEFAULT_MAPPING = {
    "surname": ["фамилия", "surname", "last name", "lastname"],
    "given_name": ["имя", "name", "first name", "firstname", "given name"],
    "latin_surname": ["фамилия лат", "latin surname"],
    "latin_given_name": ["имя лат", "latin name"],
    "birth_date": ["дата рождения", "др", "dob", "birth date", "birthdate"],
    "gender": ["пол", "gender", "sex"],
    "citizenship": ["гражданство", "citizenship", "nationality"],
    "document_number": ["паспорт", "документ", "passport", "document"],
    "document_expires": ["срок действия", "expiry", "expires"],
    "phone": ["телефон", "phone"],
    "email": ["email", "почта", "e-mail"],
    "baggage": ["багаж", "baggage"],
}

TRANSLIT = str.maketrans({
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh",
    "з": "z", "и": "i", "й": "i", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o",
    "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f", "х": "kh", "ц": "ts",
    "ч": "ch", "ш": "sh", "щ": "shch", "ъ": "", "ы": "y", "ь": "", "э": "e",
    "ю": "iu", "я": "ia",
})


def transliterate(text: str) -> str:
    """Транслитерация для экспорта; исходные данные не перезаписываются (ТЗ §11)."""
    return text.lower().translate(TRANSLIT).upper()


def parse_file(content: bytes, filename: str, column_mapping: dict | None = None) -> tuple[list, list]:
    """Возвращает (raw_rows, parsed_rows). Поддержка CSV и XLSX."""
    if filename.lower().endswith(".xlsx"):
        rows = _read_xlsx(content)
    else:
        rows = _read_csv(content)
    if not rows:
        return [], []
    headers = [str(h or "").strip().lower() for h in rows[0]]
    mapping = _resolve_mapping(headers, column_mapping)
    raw_rows = [dict(zip(headers, [str(c or "").strip() for c in row])) for row in rows[1:]]
    parsed = [_normalize_row(raw, mapping) for raw in raw_rows]
    return raw_rows, parsed


def _read_csv(content: bytes) -> list[list]:
    text = content.decode("utf-8-sig", errors="replace")
    dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t") if text.strip() else csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect) if any(row)]


def _read_xlsx(content: bytes) -> list[list]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)
            if any(cell is not None for cell in row)]


def _resolve_mapping(headers: list[str], custom: dict | None) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for field, aliases in DEFAULT_MAPPING.items():
        if custom and field in custom:
            if custom[field] in headers:
                mapping[field] = custom[field]
            continue
        for alias in aliases:
            if alias in headers:
                mapping[field] = alias
                break
    return mapping


def _normalize_row(raw: dict, mapping: dict[str, str]) -> dict:
    row = {field: raw.get(column, "") for field, column in mapping.items()}
    row["birth_date"] = _parse_date(row.get("birth_date", ""))
    row["document_expires"] = _parse_date(row.get("document_expires", ""))
    gender = str(row.get("gender", "")).strip().lower()
    row["gender"] = ("male" if gender in ("м", "муж", "m", "male")
                     else "female" if gender in ("ж", "жен", "f", "female") else "")
    row["citizenship"] = str(row.get("citizenship", "")).strip().upper()[:2]
    return row


def _parse_date(value) -> str | None:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    value = str(value or "").strip()
    if not value:
        return None
    for pattern in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(value, pattern).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def validate_row(row: dict) -> list[str]:
    """Валидация ФИО, пола, DOB, citizenship, документа (ТЗ §11)."""
    errors = []
    if not row.get("surname") and not row.get("latin_surname"):
        errors.append("Не указана фамилия")
    if not row.get("given_name") and not row.get("latin_given_name"):
        errors.append("Не указано имя")
    if not row.get("birth_date"):
        errors.append("Не указана/нераспознана дата рождения")
    if not row.get("gender"):
        errors.append("Не указан пол")
    if not row.get("citizenship"):
        errors.append("Не указано гражданство")
    if not row.get("document_number"):
        errors.append("Не указан документ")
    elif not re.fullmatch(r"[A-Za-zА-Яа-я0-9 -]{4,20}", row["document_number"]):
        errors.append("Подозрительный номер документа")
    if row.get("document_expires") is None and row.get("document_number"):
        errors.append("Не указан срок действия документа")
    return errors


def build_preview(order, parsed_rows: list[dict]) -> dict:
    """Сравнение с участниками заказа: same/changed/new/missing/conflict."""
    participants = list(
        order.participants.filter(status="active").select_related("person")
    )
    matched_participants: set = set()
    items = []
    for index, row in enumerate(parsed_rows):
        errors = validate_row(row)
        match, conflicts = _match_participant(row, participants)
        if match is not None:
            matched_participants.add(match.pk)
            state = ("conflict" if conflicts
                     else "changed" if _differs(row, match.person) else "same")
        else:
            state = "new"
        items.append({
            "row_index": index, "row": row, "state": state,
            "participant_id": str(match.pk) if match else None,
            "conflicts": conflicts, "validation_errors": errors,
        })
    missing = [
        {"participant_id": str(p.pk),
         "name": p.person.full_name if p.person else (p.guest_snapshot or {}).get("surname", "")}
        for p in participants if p.pk not in matched_participants
    ]
    return {"items": items, "missing": missing,
            "stats": _stats(items, missing)}


def _stats(items: list, missing: list) -> dict:
    stats = {"same": 0, "changed": 0, "new": 0, "conflict": 0, "invalid": 0}
    for item in items:
        stats[item["state"]] = stats.get(item["state"], 0) + 1
        if item["validation_errors"]:
            stats["invalid"] += 1
    stats["missing"] = len(missing)
    return stats


def _match_participant(row: dict, participants: list):
    surname = (row.get("surname") or row.get("latin_surname", "")).lower()
    given = (row.get("given_name") or row.get("latin_given_name", "")).lower()
    for participant in participants:
        person = participant.person
        if person is None:
            continue
        person_surnames = {person.surname.lower(), person.latin_surname.lower()}
        person_names = {person.given_name.lower(), person.latin_given_name.lower()}
        if surname in person_surnames and given in person_names:
            conflicts = []
            if row.get("birth_date") and person.birth_date and \
                    str(person.birth_date) != row["birth_date"]:
                conflicts.append({"field": "birth_date", "current": str(person.birth_date),
                                  "incoming": row["birth_date"]})
            return participant, conflicts
    return None, []


def _differs(row: dict, person) -> bool:
    if person is None:
        return False
    checks = [
        ("phone", person.phone), ("email", person.email),
        ("gender", person.gender), ("citizenship", person.citizenship),
    ]
    for field, current in checks:
        incoming = row.get(field)
        if incoming and str(incoming) != str(current):
            return True
    if row.get("birth_date") and person.birth_date and \
            row["birth_date"] != str(person.birth_date):
        return True
    return False
