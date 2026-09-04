import json

from rest_framework import serializers
from rest_framework import status as http
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import has_permission, require
from common.audit import audit
from common.errors import ApiError
from crm.models import Company
from travel_policy.models import TravelPolicy, check_offer_compliance


class TravelPolicySerializer(serializers.ModelSerializer):
    updated_by_name = serializers.SerializerMethodField()

    class Meta:
        model = TravelPolicy
        fields = [
            "id",
            "company",
            "name",
            "effective_from",
            "effective_to",
            "is_active",
            "policy_version",
            "scopes",
            "allowed_avia_cabins",
            "allowed_airlines",
            "allowed_rail_classes",
            "allowed_train_types",
            "allowed_hotel_categories",
            "allowed_hotel_chains",
            "allowed_meal_plans",
            "allowed_car_classes",
            "price_limits",
            "min_advance_booking_days",
            "approver_chain",
            "created_at",
            "updated_at",
            "updated_by_name",
        ]
        read_only_fields = ["id", "policy_version", "company", "created_at", "updated_at", "updated_by_name"]

    def get_updated_by_name(self, obj):
        user = obj.updated_by or obj.created_by
        return user.get_full_name() if user else ""


def _as_list(value):
    if isinstance(value, list):
        return value
    if value in (None, ""):
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def _ui_policy_payload(ui):
    avia = ui.get("avia") or {}
    rail = ui.get("rail") or {}
    hotels = ui.get("hotels") or {}
    transfers = ui.get("transfers") or {}
    approval = ui.get("approval") or {}

    def limit(section, amount_key, currency_key):
        amount = section.get(amount_key)
        return {"amount": amount, "currency": section.get(currency_key) or "RUB"} if amount not in (None, "") else None

    limits = {
        "avia": limit(avia, "maxPrice", "maxPriceCur"),
        "rail": limit(rail, "maxPrice", "maxPriceCur"),
        "hotel": limit(hotels, "maxNight", "maxNightCur"),
        "transfer": limit(transfers, "maxPrice", "maxPriceCur"),
    }
    return {
        "name": ui.get("name") or "Тревел-политика компании",
        "is_active": True,
        "scopes": [{"scope": ui.get("scope") or "Вся компания", "value": ui.get("scopeValue") or "", "ui": ui}],
        "allowed_avia_cabins": _as_list(avia.get("classAllowed")),
        "allowed_airlines": _as_list(avia.get("airlinesAllowed")),
        "allowed_rail_classes": _as_list(rail.get("wagonClass")),
        "allowed_train_types": _as_list(rail.get("wagonTypes")),
        "allowed_hotel_categories": _as_list(hotels.get("maxCategory")),
        "allowed_hotel_chains": _as_list(hotels.get("chainsAllowed")),
        "allowed_meal_plans": _as_list(hotels.get("boardAllowed")),
        "allowed_car_classes": _as_list(transfers.get("carClasses")),
        "price_limits": {key: value for key, value in limits.items() if value},
        "min_advance_booking_days": avia.get("minLeadDays") or rail.get("minLeadDays") or None,
        "approver_chain": _as_list(approval.get("approvers")),
    }


def _policy_upload_payload(upload):
    if upload.size > 10 * 1024 * 1024:
        raise ApiError(code="FILE_TOO_LARGE", message="Файл больше 10 МБ", status_code=400)
    filename = (upload.name or "").lower()
    if filename.endswith(".json"):
        try:
            payload = json.loads(upload.read().decode("utf-8-sig"))
        except (UnicodeDecodeError, json.JSONDecodeError) as error:
            raise ApiError(code="INVALID_FILE", message="Некорректный JSON", status_code=400) from error
    elif filename.endswith(".xlsx"):
        from openpyxl import load_workbook

        try:
            rows = list(load_workbook(upload, read_only=True, data_only=True).active.iter_rows(values_only=True))
        except Exception as error:
            raise ApiError(code="INVALID_FILE", message="Не удалось прочитать XLSX", status_code=400) from error
        payload = {}
        for row in rows:
            if len(row) < 2 or not row[0]:
                continue
            key, value = str(row[0]).strip(), row[1]
            if isinstance(value, str) and value[:1] in "[{":
                try:
                    value = json.loads(value)
                except json.JSONDecodeError:
                    pass
            payload[key] = value
    else:
        raise ApiError(code="UNSUPPORTED_FILE", message="Поддерживаются только JSON и XLSX", status_code=400)
    if not isinstance(payload, dict):
        raise ApiError(code="INVALID_FILE", message="Документ должен содержать объект политики", status_code=400)
    if isinstance(payload.get("policy"), dict):
        return _ui_policy_payload(payload["policy"])
    if any(key in payload for key in ("avia", "rail", "hotels", "transfers", "approval")):
        return _ui_policy_payload(payload)
    return payload


class CompanyTravelPoliciesView(APIView):
    permission_classes = [require("crm.view")]

    def get(self, request, company_id):
        policies = TravelPolicy.objects.filter(
            company_id=company_id, tenant_id=request.user.tenant_id, archived_at__isnull=True
        )
        return Response(TravelPolicySerializer(policies, many=True).data)

    def post(self, request, company_id):
        if not has_permission(request.user, "crm.change"):
            raise ApiError(code="PERMISSION_DENIED", message="Нет права crm.change", status_code=403)
        company = Company.objects.filter(pk=company_id, tenant_id=request.user.tenant_id).first()
        if company is None:
            raise ApiError(code="NOT_FOUND", message="Компания не найдена", status_code=404)
        serializer = TravelPolicySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        policy = serializer.save(tenant_id=request.user.tenant_id, company=company, created_by=request.user)
        audit("travel_policy.created", actor=request.user, resource=policy, request=request)
        return Response(TravelPolicySerializer(policy).data, status=http.HTTP_201_CREATED)


class CompanyTravelPolicyImportView(APIView):
    permission_classes = [require("crm.change")]

    def post(self, request, company_id):
        company = Company.objects.filter(pk=company_id, tenant_id=request.user.tenant_id).first()
        if company is None:
            raise ApiError(code="NOT_FOUND", message="Компания не найдена", status_code=404)
        upload = request.FILES.get("file")
        if upload is None:
            raise ApiError(code="VALIDATION_ERROR", message="Выберите файл", status_code=400)
        payload = _policy_upload_payload(upload)
        policy = TravelPolicy.objects.filter(
            company=company,
            tenant_id=request.user.tenant_id,
            is_active=True,
            archived_at__isnull=True,
        ).order_by("-updated_at").first()
        serializer = TravelPolicySerializer(policy, data=payload, partial=policy is not None)
        serializer.is_valid(raise_exception=True)
        if policy is None:
            policy = serializer.save(tenant_id=request.user.tenant_id, company=company, created_by=request.user)
            action = "travel_policy.imported"
        else:
            policy = serializer.save(
                policy_version=policy.policy_version + 1,
                updated_by=request.user,
            )
            action = "travel_policy.imported_update"
        audit(action, actor=request.user, resource=policy, request=request)
        return Response(TravelPolicySerializer(policy).data, status=http.HTTP_201_CREATED)


class TravelPolicyDetailView(APIView):
    permission_classes = [require("crm.view")]

    def _get(self, request, policy_id) -> TravelPolicy:
        policy = TravelPolicy.objects.filter(pk=policy_id, tenant_id=request.user.tenant_id).first()
        if policy is None:
            raise ApiError(code="NOT_FOUND", message="Политика не найдена", status_code=404)
        return policy

    def get(self, request, policy_id):
        return Response(TravelPolicySerializer(self._get(request, policy_id)).data)

    def patch(self, request, policy_id):
        if not has_permission(request.user, "crm.change"):
            raise ApiError(code="PERMISSION_DENIED", message="Нет права crm.change", status_code=403)
        policy = self._get(request, policy_id)
        serializer = TravelPolicySerializer(policy, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save(policy_version=policy.policy_version + 1, updated_by=request.user)
        audit("travel_policy.updated", actor=request.user, resource=policy, request=request)
        return Response(serializer.data)


class TravelPolicyCheckView(APIView):
    """POST /travel-policies/{id}/check/ — compliance оффера (ТЗ §6.3)."""

    permission_classes = [require("crm.view")]

    def post(self, request, policy_id):
        policy = TravelPolicy.objects.filter(pk=policy_id, tenant_id=request.user.tenant_id).first()
        if policy is None:
            raise ApiError(code="NOT_FOUND", message="Политика не найдена", status_code=404)
        offer = request.data.get("offer")
        if not isinstance(offer, dict):
            raise ApiError(
                code="VALIDATION_ERROR",
                message="Ожидается объект offer",
                fields={"offer": ["Обязательное поле-объект"]},
                status_code=400,
            )
        result = check_offer_compliance(policy, offer)
        return Response(result.as_dict(policy))
